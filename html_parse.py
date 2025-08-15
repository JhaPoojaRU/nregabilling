# To access all HTML files
import glob
# For HTML pARSING
from bs4 import BeautifulSoup
# To use excel to extract data
from openpyxl import Workbook
from openpyxl import load_workbook
# For cast typing date into excel
from datetime import datetime

# Uses classes created as Lib
from billclass import BillData, MaterialData, VendorData


html_files = glob.glob('reports/*.html')
excel_file = "data/vendor_bill_data.xlsx"

def parse_html():
      for html_file in html_files:
            print("------------------NREGA DATA BY BLOCK AND PANCHAYAT------------------")
            print("Current File Running",html_file)
            with open(html_file, 'r', encoding="utf-8") as f:
                  soup = BeautifulSoup(f, 'html.parser')
                  tables = soup.find_all('table')
                  len(tables)

                  block_table = tables[3]
                  block_info = block_table.find_all('td')
            
                  print(block_info[2].get_text())
                  print(block_info[3].get_text())

                  # finding the block and panchayat name because there will be multiple files with different names later
                  block = block_info[2].get_text().replace("Block", "").replace(":", "").strip()
                  panchayat = block_info[3].get_text().replace("Panchayat", "").replace(":", "").strip()
                  # Only using tables that are required for bill related data
                  data_tables = tables[5:]

                  data = []

                  vendor_names = ['navita enterprises', 'mithila tool kits hardware and nursery', 'navita enterprises and nursery', 'pallavi enterprises and nursery']

                  for table in data_tables:
                        rows = table.find_all('tr')

                        i = 0
                        while( i < len(rows) ):

                              if any(vendor_name in rows[i].get_text().lower() for vendor_name in vendor_names):
                                    vendor_name = next(vendor_name for vendor_name in vendor_names if vendor_name in rows[i].get_text().lower())
                                    bill_row           = rows[i-1]
                                    vendor_row         = rows[i]
                                    material_row       = rows[i+1]
                                    material_value_row = rows[i+2]

                                    ven_data = VendorData(vendor_name, vendor_row, bill_row, material_value_row)

                                    vendor_data = ven_data.get_data_dict()

                              

                                    data.append(vendor_data)


                                    i = i + 3
                                    continue
                              else:
                                    i = i + 1

                  if data:
                        print("Data found")
                        print("Inserting", len(data),"rows to excel")

                        try:
                              wb = load_workbook(excel_file)
                              ws = wb.active
                        except FileNotFoundError:
                              wb = Workbook()
                              ws = wb.active
                              ws.append(["block", "panchayat", "vendor_name", "bill_no", "bill_amt",
                                          "bill_dt", "bill_dop", "fin_year", "material", "unit_price",
                                          "qty", "amt"])

                        for vendor_data in data:

                              # Convert numeric data to correct data types
                              try:
                                bill_no = int(vendor_data["bill_no"]) if vendor_data["bill_no"] else None
                              except ValueError:
                                bill_no = None

                              try:
                                bill_amount = float(vendor_data["bill_amt"]) if vendor_data["bill_amt"] else None
                              except ValueError:
                                bill_amount = None

                              try:
                                unit_price = float(vendor_data["unit_price"]) if vendor_data["unit_price"] else None
                              except ValueError:
                                unit_price = None

                              try:
                                quantity = float(vendor_data["quantity"]) if vendor_data["quantity"] else None
                              except ValueError:
                                quantity = None

                              try:
                                final_amt = float(vendor_data["amount"]) if vendor_data["amount"] else None
                              except ValueError:
                                final_amt = None

                              # Convert date strings to datetime objects
                              try:
                                bill_date = datetime.strptime(vendor_data["bill_date"], '%d/%m/%Y').date() if vendor_data["bill_date"] else None
                              except ValueError:
                                bill_date = None

                              try:
                                bill_dop = datetime.strptime(vendor_data["bill_dop"], '%d/%m/%Y').date() if vendor_data["bill_dop"] else None
                              except ValueError:
                                bill_dop = None  

                              ws.append([block, panchayat, vendor_data["vendor_name"], bill_no,
                                          bill_amount, bill_date, bill_dop, vendor_data["fin_year"],
                                          vendor_data["material"], unit_price, quantity, final_amt])
                        wb.save(excel_file)
                  else:
                        print("No data found")


if __name__ == "__main__":
      parse_html()
